from __future__ import annotations

import re
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


APP_DIR = Path(__file__).parent
DB_PATH = APP_DIR / "data" / "instruments.db"
UPLOAD_DIR = APP_DIR / "uploads"
OUTPUT_DIR = APP_DIR / "outputs"
DEFAULT_EXCEL_PATH = Path(r"C:\Users\김상훈\Desktop\계측기 등록 대장_2026.05.15.xlsx")

DATE_PATTERN = re.compile(r"(\d{2,4})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{1,2})")


@dataclass
class ImportSummary:
    instrument_count: int
    record_count: int
    disposed_count: int
    source_path: str


SCHEMA = """
CREATE TABLE IF NOT EXISTS instruments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    management_no TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    serial_no TEXT,
    cycle_text TEXT,
    cycle_months INTEGER,
    location TEXT,
    process TEXT,
    department TEXT,
    department_owner TEXT,
    department_owner2 TEXT,
    qc_owner TEXT,
    is_standard INTEGER DEFAULT 0,
    status TEXT DEFAULT '사용',
    remark TEXT,
    history_card_updated TEXT,
    correction_offset REAL DEFAULT 0,
    correction_factor REAL DEFAULT 1,
    correction_unit TEXT,
    correction_note TEXT,
    disposal_report_no TEXT,
    disposal_report_file_path TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS calibration_records (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    instrument_id INTEGER NOT NULL,
    calibration_type TEXT NOT NULL,
    calibration_date TEXT,
    next_due_date TEXT,
    result TEXT,
    certificate_no TEXT,
    certificate_file_path TEXT,
    measured_value REAL,
    corrected_value REAL,
    correction_snapshot TEXT,
    note TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY (instrument_id) REFERENCES instruments(id)
);

CREATE TABLE IF NOT EXISTS department_contacts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    department TEXT UNIQUE NOT NULL,
    owner_name TEXT,
    kakao_target TEXT,
    phone TEXT,
    note TEXT,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS import_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_path TEXT,
    imported_at TEXT NOT NULL,
    instrument_count INTEGER,
    record_count INTEGER
);
"""


def connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def ensure_columns(conn: sqlite3.Connection, table: str, columns: dict[str, str]) -> None:
    existing = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    for name, ddl in columns.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {name} {ddl}")


def init_db() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        conn.executescript(SCHEMA)
        ensure_columns(
            conn,
            "instruments",
            {
                "process": "TEXT",
                "department_owner2": "TEXT",
                "disposal_report_no": "TEXT",
                "disposal_report_file_path": "TEXT",
            },
        )
        ensure_columns(conn, "calibration_records", {"certificate_file_path": "TEXT"})


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def parse_cycle_months(text: str) -> int | None:
    match = re.search(r"(\d+)\s*개월", text or "")
    return int(match.group(1)) if match else None


def parse_dates(text: Any) -> list[date]:
    dates: list[date] = []
    for y, m, d in DATE_PATTERN.findall(clean_text(text)):
        year = int(y)
        if year < 100:
            year += 2000
        try:
            dates.append(date(year, int(m), int(d)))
        except ValueError:
            pass
    return dates


def derive_department(location: str) -> str:
    loc = clean_text(location) or "미지정"
    if loc.startswith("QC"):
        return "QC"
    if "사출" in loc:
        return "사출"
    if "출하" in loc:
        return "출하"
    if "원료" in loc:
        return "원료생산팀"
    if "하드" in loc:
        return "하드실"
    if "멸균" in loc:
        return "멸균"
    if "생기" in loc or "보전" in loc:
        return "생기보전"
    if "최종" in loc:
        return "최종검사"
    return loc.split()[0]


def latest_record_df() -> pd.DataFrame:
    query = """
    WITH ranked AS (
        SELECT
            cr.*,
            ROW_NUMBER() OVER (
                PARTITION BY cr.instrument_id
                ORDER BY
                    CASE WHEN cr.next_due_date IS NULL OR cr.next_due_date = '' THEN 1 ELSE 0 END,
                    cr.next_due_date DESC,
                    cr.id DESC
            ) AS rn
        FROM calibration_records cr
    )
    SELECT
        i.*,
        r.id AS last_record_id,
        r.calibration_type AS last_calibration_type,
        r.calibration_date AS last_calibration_date,
        r.next_due_date,
        r.result AS last_result,
        r.certificate_no AS last_certificate_no,
        r.certificate_file_path AS last_certificate_file_path,
        r.measured_value AS last_measured_value,
        r.corrected_value AS last_corrected_value
    FROM instruments i
    LEFT JOIN ranked r ON r.instrument_id = i.id AND r.rn = 1
    """
    with connect() as conn:
        return pd.read_sql_query(query, conn)


def instruments_df(include_disposed: bool = True) -> pd.DataFrame:
    df = latest_record_df()
    if not include_disposed and not df.empty:
        df = df[df["status"] != "폐기"]
    return df


def calibration_history_df(instrument_id: int | None = None) -> pd.DataFrame:
    where = ""
    params: tuple[Any, ...] = ()
    if instrument_id:
        where = "WHERE cr.instrument_id = ?"
        params = (instrument_id,)
    query = f"""
    SELECT
        cr.id,
        cr.instrument_id,
        i.management_no,
        i.name,
        cr.calibration_type,
        cr.calibration_date,
        cr.next_due_date,
        cr.result,
        cr.certificate_no,
        cr.certificate_file_path,
        cr.measured_value,
        cr.corrected_value,
        cr.note,
        cr.created_at
    FROM calibration_records cr
    JOIN instruments i ON i.id = cr.instrument_id
    {where}
    ORDER BY cr.next_due_date DESC, cr.id DESC
    """
    with connect() as conn:
        return pd.read_sql_query(query, conn, params=params)


def contacts_df() -> pd.DataFrame:
    with connect() as conn:
        return pd.read_sql_query("SELECT * FROM department_contacts ORDER BY department", conn)


def upsert_contact(department: str, owner_name: str, kakao_target: str, phone: str, note: str) -> None:
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO department_contacts (department, owner_name, kakao_target, phone, note, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(department) DO UPDATE SET
                owner_name=excluded.owner_name,
                kakao_target=excluded.kakao_target,
                phone=excluded.phone,
                note=excluded.note,
                updated_at=excluded.updated_at
            """,
            (department, owner_name, kakao_target, phone, note, now_text()),
        )


def upsert_instrument(data: dict[str, Any]) -> int:
    ts = now_text()
    cycle_text = clean_text(data.get("cycle_text"))
    payload = {
        "management_no": clean_text(data["management_no"]),
        "name": clean_text(data["name"]),
        "serial_no": clean_text(data.get("serial_no")),
        "cycle_text": cycle_text,
        "cycle_months": data.get("cycle_months") or parse_cycle_months(cycle_text),
        "location": clean_text(data.get("location")),
        "process": clean_text(data.get("process")),
        "department": clean_text(data.get("department")) or derive_department(data.get("location", "")),
        "department_owner": clean_text(data.get("department_owner")),
        "department_owner2": clean_text(data.get("department_owner2")),
        "qc_owner": clean_text(data.get("qc_owner")),
        "is_standard": 1 if data.get("is_standard") else 0,
        "status": clean_text(data.get("status")) or "사용",
        "remark": clean_text(data.get("remark")),
        "history_card_updated": clean_text(data.get("history_card_updated")),
        "correction_offset": float(data.get("correction_offset") or 0),
        "correction_factor": float(data.get("correction_factor") or 1),
        "correction_unit": clean_text(data.get("correction_unit")),
        "correction_note": clean_text(data.get("correction_note")),
        "created_at": ts,
        "updated_at": ts,
    }
    with connect() as conn:
        row = conn.execute("SELECT id FROM instruments WHERE management_no = ?", (payload["management_no"],)).fetchone()
        if row:
            conn.execute(
                """
                UPDATE instruments SET
                    name=:name, serial_no=:serial_no, cycle_text=:cycle_text, cycle_months=:cycle_months,
                    location=:location, process=:process, department=:department,
                    department_owner=:department_owner, department_owner2=:department_owner2,
                    qc_owner=:qc_owner, is_standard=:is_standard, status=:status, remark=:remark,
                    history_card_updated=:history_card_updated, correction_offset=:correction_offset,
                    correction_factor=:correction_factor, correction_unit=:correction_unit,
                    correction_note=:correction_note, updated_at=:updated_at
                WHERE management_no=:management_no
                """,
                payload,
            )
            return int(row["id"])
        conn.execute(
            """
            INSERT INTO instruments (
                management_no, name, serial_no, cycle_text, cycle_months, location, process, department,
                department_owner, department_owner2, qc_owner, is_standard, status, remark, history_card_updated,
                correction_offset, correction_factor, correction_unit, correction_note, created_at, updated_at
            ) VALUES (
                :management_no, :name, :serial_no, :cycle_text, :cycle_months, :location, :process, :department,
                :department_owner, :department_owner2, :qc_owner, :is_standard, :status, :remark, :history_card_updated,
                :correction_offset, :correction_factor, :correction_unit, :correction_note, :created_at, :updated_at
            )
            """,
            payload,
        )
        return int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])


def add_calibration_record(data: dict[str, Any]) -> None:
    with connect() as conn:
        payload = {
            "instrument_id": data["instrument_id"],
            "calibration_type": clean_text(data.get("calibration_type")),
            "calibration_date": clean_text(data.get("calibration_date")),
            "next_due_date": clean_text(data.get("next_due_date")),
            "result": clean_text(data.get("result")),
            "certificate_no": clean_text(data.get("certificate_no")),
            "certificate_file_path": clean_text(data.get("certificate_file_path")),
            "measured_value": data.get("measured_value"),
            "corrected_value": data.get("corrected_value"),
            "correction_snapshot": clean_text(data.get("correction_snapshot")),
            "note": clean_text(data.get("note")),
        }
        existing = conn.execute(
            """
            SELECT id, certificate_no, certificate_file_path, measured_value, corrected_value, correction_snapshot
            FROM calibration_records
            WHERE instrument_id = ? AND calibration_type = ? AND calibration_date = ?
              AND next_due_date = ? AND COALESCE(note, '') = ?
            """,
            (
                payload["instrument_id"],
                payload["calibration_type"],
                payload["calibration_date"],
                payload["next_due_date"],
                payload["note"],
            ),
        ).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE calibration_records
                SET result = COALESCE(NULLIF(?, ''), result),
                    certificate_no = COALESCE(NULLIF(?, ''), certificate_no),
                    certificate_file_path = COALESCE(NULLIF(?, ''), certificate_file_path),
                    measured_value = COALESCE(?, measured_value),
                    corrected_value = COALESCE(?, corrected_value),
                    correction_snapshot = COALESCE(NULLIF(?, ''), correction_snapshot)
                WHERE id = ?
                """,
                (
                    payload["result"],
                    payload["certificate_no"],
                    payload["certificate_file_path"],
                    payload["measured_value"],
                    payload["corrected_value"],
                    payload["correction_snapshot"],
                    existing["id"],
                ),
            )
            return
        conn.execute(
            """
            INSERT INTO calibration_records (
                instrument_id, calibration_type, calibration_date, next_due_date, result, certificate_no,
                certificate_file_path, measured_value, corrected_value, correction_snapshot, note, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload["instrument_id"],
                payload["calibration_type"],
                payload["calibration_date"],
                payload["next_due_date"],
                payload["result"],
                payload["certificate_no"],
                payload["certificate_file_path"],
                payload["measured_value"],
                payload["corrected_value"],
                payload["correction_snapshot"],
                payload["note"],
                now_text(),
            ),
        )


def update_instrument_master(instrument_id: int, data: dict[str, Any]) -> None:
    with connect() as conn:
        conn.execute(
            """
            UPDATE instruments
            SET name = ?, serial_no = ?, cycle_text = ?, cycle_months = ?, location = ?, process = ?,
                department = ?, department_owner = ?, department_owner2 = ?, qc_owner = ?,
                is_standard = ?, status = ?, remark = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                clean_text(data.get("name")),
                clean_text(data.get("serial_no")),
                clean_text(data.get("cycle_text")),
                data.get("cycle_months") or parse_cycle_months(clean_text(data.get("cycle_text"))) or None,
                clean_text(data.get("location")),
                clean_text(data.get("process")),
                clean_text(data.get("department")),
                clean_text(data.get("department_owner")),
                clean_text(data.get("department_owner2")),
                clean_text(data.get("qc_owner")),
                1 if data.get("is_standard") else 0,
                clean_text(data.get("status")) or "사용",
                clean_text(data.get("remark")),
                now_text(),
                instrument_id,
            ),
        )


def update_correction(instrument_id: int, offset: float, factor: float, unit: str, note: str) -> None:
    with connect() as conn:
        conn.execute(
            """
            UPDATE instruments
            SET correction_offset = ?, correction_factor = ?, correction_unit = ?, correction_note = ?, updated_at = ?
            WHERE id = ?
            """,
            (offset, factor, clean_text(unit), clean_text(note), now_text(), instrument_id),
        )


def mark_disposed(instrument_id: int, note: str, report_no: str = "", report_file_path: str = "") -> None:
    report_note = clean_text(note)
    with connect() as conn:
        conn.execute(
            """
            UPDATE instruments
            SET status = '폐기',
                disposal_report_no = COALESCE(NULLIF(?, ''), disposal_report_no),
                disposal_report_file_path = COALESCE(NULLIF(?, ''), disposal_report_file_path),
                remark = CASE
                    WHEN ? = '' THEN remark
                    WHEN remark IS NULL OR remark = '' THEN ?
                    ELSE remark || '\n' || ?
                END,
                updated_at = ?
            WHERE id = ?
            """,
            (clean_text(report_no), clean_text(report_file_path), report_note, report_note, report_note, now_text(), instrument_id),
        )


def calculate_corrected(measured: float, offset: float, factor: float) -> float:
    return (measured + offset) * factor


def get_import_log() -> pd.DataFrame:
    init_db()
    with connect() as conn:
        return pd.read_sql_query("SELECT * FROM import_log ORDER BY id DESC", conn)


def get_instrument(instrument_id: int) -> dict[str, Any] | None:
    with connect() as conn:
        row = conn.execute("SELECT * FROM instruments WHERE id = ?", (instrument_id,)).fetchone()
        return dict(row) if row else None


def import_excel(path: str | Path, reset: bool = False) -> ImportSummary:
    init_db()
    source = Path(path)
    wb = openpyxl.load_workbook(source, data_only=True)
    instrument_count = 0
    record_count = 0
    disposed_count = 0

    with connect() as conn:
        if reset:
            conn.execute("DELETE FROM calibration_records")
            conn.execute("DELETE FROM instruments")

    ws = wb["검교정 내역-2026.05.15"] if "검교정 내역-2026.05.15" in wb.sheetnames else wb.worksheets[0]
    for row in ws.iter_rows(min_row=4, values_only=True):
        management_no = clean_text(row[0] if len(row) > 0 else "")
        name = clean_text(row[1] if len(row) > 1 else "")
        if not management_no or not name:
            continue
        remark = clean_text(row[12] if len(row) > 12 else "")
        disposal_text = " ".join([remark, clean_text(row[9] if len(row) > 9 else ""), clean_text(row[10] if len(row) > 10 else "")])
        status = "폐기" if "폐기" in disposal_text else "사용"
        disposed_count += 1 if status == "폐기" else 0
        instrument_id = upsert_instrument(
            {
                "management_no": management_no,
                "name": name,
                "serial_no": clean_text(row[2] if len(row) > 2 else ""),
                "cycle_text": clean_text(row[3] if len(row) > 3 else ""),
                "location": clean_text(row[4] if len(row) > 4 else ""),
                "process": clean_text(row[5] if len(row) > 5 else ""),
                "department": clean_text(row[6] if len(row) > 6 else "") or derive_department(row[4] if len(row) > 4 else ""),
                "department_owner": clean_text(row[7] if len(row) > 7 else ""),
                "department_owner2": clean_text(row[8] if len(row) > 8 else ""),
                "is_standard": "표준품" in management_no or "표준품" in remark,
                "status": status,
                "remark": remark,
                "history_card_updated": clean_text(row[13] if len(row) > 13 else ""),
            }
        )
        instrument_count += 1
        for calibration_type, value in (("내부", row[9] if len(row) > 9 else None), ("외부", row[10] if len(row) > 10 else None)):
            text = clean_text(value)
            if not text or text.upper() == "N/A" or "폐기" in text:
                continue
            parsed = parse_dates(text)
            add_calibration_record(
                {
                    "instrument_id": instrument_id,
                    "calibration_type": calibration_type,
                    "calibration_date": parsed[0].isoformat() if parsed else "",
                    "next_due_date": parsed[-1].isoformat() if len(parsed) > 1 else "",
                    "result": "기존 대장",
                    "note": text,
                }
            )
            record_count += 1

    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            sterilizer_no = clean_text(row[0] if len(row) > 0 else "")
            name = clean_text(row[1] if len(row) > 1 else "")
            serial_no = clean_text(row[2] if len(row) > 2 else "")
            sensor_no = clean_text(row[3] if len(row) > 3 else "")
            if not name or not serial_no:
                continue
            instrument_id = upsert_instrument(
                {
                    "management_no": f"멸균-{sterilizer_no}-{serial_no}".replace(" ", ""),
                    "name": name,
                    "serial_no": serial_no,
                    "cycle_text": "12개월",
                    "location": sterilizer_no,
                    "process": "멸균기 부착",
                    "department": "멸균",
                    "status": "사용",
                    "remark": f"멸균기 센서 No: {sensor_no}" if sensor_no else "",
                }
            )
            instrument_count += 1
            cal_dates = parse_dates(row[4] if len(row) > 4 else "")
            due_dates = parse_dates(row[5] if len(row) > 5 else "")
            add_calibration_record(
                {
                    "instrument_id": instrument_id,
                    "calibration_type": "외부",
                    "calibration_date": cal_dates[0].isoformat() if cal_dates else "",
                    "next_due_date": due_dates[0].isoformat() if due_dates else "",
                    "result": "기존 대장",
                    "note": "멸균기 부착 계측기",
                }
            )
            record_count += 1

    with connect() as conn:
        conn.execute(
            "INSERT INTO import_log (source_path, imported_at, instrument_count, record_count) VALUES (?, ?, ?, ?)",
            (str(source), now_text(), instrument_count, record_count),
        )
    return ImportSummary(instrument_count, record_count, disposed_count, str(source))


def dashboard_metrics() -> dict[str, int]:
    df = instruments_df(include_disposed=True)
    if df.empty:
        return {"total": 0, "active": 0, "disposed": 0, "overdue": 0, "due_90": 0}
    today = pd.Timestamp(date.today())
    due = pd.to_datetime(df["next_due_date"], errors="coerce")
    active = df["status"] != "폐기"
    return {
        "total": int(len(df)),
        "active": int(active.sum()),
        "disposed": int((df["status"] == "폐기").sum()),
        "overdue": int(((due < today) & active).sum()),
        "due_90": int(((due >= today) & (due <= today + pd.Timedelta(days=90)) & active).sum()),
    }


def due_items(days: int = 90, include_overdue: bool = True, calibration_filter: str = "전체") -> pd.DataFrame:
    df = instruments_df(include_disposed=False)
    if df.empty:
        return df
    if calibration_filter != "전체":
        df = df[df["last_calibration_type"].fillna("") == calibration_filter]
    today = pd.Timestamp(date.today())
    due = pd.to_datetime(df["next_due_date"], errors="coerce")
    mask = due <= today + pd.Timedelta(days=days)
    if not include_overdue:
        mask &= due >= today
    df = df[mask].copy()
    df["due_dt"] = due[mask]
    df["남은일수"] = (df["due_dt"] - today).dt.days
    return df.sort_values(["남은일수", "department", "management_no"])


def make_kakao_message(department: str, rows: pd.DataFrame, owner: str = "") -> str:
    today = date.today().isoformat()
    target = "전체 사용부서" if department == "전체" else department
    title_owner = f"{owner}님, " if owner and department != "전체" else ""
    lines = [
        f"{title_owner}검교정 도래 계측기 반납 요청드립니다.",
        f"대상: {target}",
        f"기준일: {today}",
        "",
    ]
    for _, row in rows.iterrows():
        due = row.get("next_due_date") or "일자 미입력"
        days = row.get("남은일수")
        days_text = f"D{int(days):+d}" if pd.notna(days) else "D-?"
        process = f" / 공정: {row.get('process')}" if row.get("process") else ""
        owner1 = row.get("department_owner") or ""
        owner2 = row.get("department_owner2") or ""
        owners = ", ".join([x for x in [owner1, owner2] if x and x != "N/A"])
        owner_text = f" / 담당: {owners}" if owners else ""
        lines.append(
            f"- [{row.get('department','')}] {row['management_no']} / {row['name']} / {due} ({days_text})"
            f" / 위치: {row.get('location','')}{process}{owner_text}"
        )
    lines += ["", "교정 예정일 전 QC로 반납 부탁드립니다.", "이미 반납 또는 교정 진행 중이면 회신 부탁드립니다."]
    return "\n".join(lines)


def dataframe_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "조회결과") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)
    return output.getvalue()


def safe_name(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣_.-]+", "_", clean_text(value)).strip("_") or "file"


def save_uploaded_file(uploaded_file: Any, category: str, management_no: str) -> str:
    if uploaded_file is None:
        return ""
    target_dir = UPLOAD_DIR / category / safe_name(management_no)
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name(uploaded_file.name)}"
    path = target_dir / filename
    with open(path, "wb") as file:
        file.write(uploaded_file.getbuffer())
    return str(path)


def guess_document_no(uploaded_file: Any) -> str:
    if uploaded_file is None:
        return ""
    stem = Path(uploaded_file.name).stem
    match = re.search(r"([A-Za-z가-힣]*[-_]?\d{2,}[-_]?\d*)", stem)
    return match.group(1) if match else ""


def export_internal_certificate(instrument: dict[str, Any], record: dict[str, Any], output_dir: Path = OUTPUT_DIR) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "내부교정 성적서"
    rows = [
        ("내부교정 성적서", "", ""),
        ("", "", ""),
        ("관리번호", instrument["management_no"], ""),
        ("계측기명", instrument["name"], ""),
        ("제작 일련번호", instrument.get("serial_no", ""), ""),
        ("설치 위치", instrument.get("location", ""), ""),
        ("사용부서", instrument.get("department", ""), ""),
        ("", "", ""),
        ("교정일자", record.get("calibration_date", ""), ""),
        ("차기교정일", record.get("next_due_date", ""), ""),
        ("측정값", record.get("measured_value", ""), instrument.get("correction_unit", "")),
        ("보정값(더하기)", instrument.get("correction_offset", 0), ""),
        ("보정계수(곱하기)", instrument.get("correction_factor", 1), ""),
        ("보정 적용값", record.get("corrected_value", ""), instrument.get("correction_unit", "")),
        ("", "", ""),
        ("판정", record.get("result", ""), ""),
        ("비고", record.get("note", ""), ""),
        ("", "", ""),
        ("보정식", "보정 적용값 = (측정값 + 보정값) x 보정계수", ""),
        ("", "", ""),
        ("작성", "QC", ""),
        ("확인", "", ""),
    ]
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx).value = value
    ws["A1"].font = openpyxl.styles.Font(size=16, bold=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 12
    for row in ws.iter_rows(min_row=3, max_row=22, min_col=1, max_col=3):
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin"),
            )
            cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
    path = output_dir / f"내부교정성적서_{safe_name(instrument['management_no'])}_{date.today().isoformat()}.xlsx"
    wb.save(path)
    return path
