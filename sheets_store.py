from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from cloud_integrations import get_secret, sheets_service
from store import (
    DEFAULT_EXCEL_PATH,
    OUTPUT_DIR,
    ImportSummary,
    calculate_corrected,
    clean_text,
    dataframe_to_xlsx_bytes,
    derive_department,
    export_internal_certificate,
    guess_document_no,
    make_kakao_message,
    parse_cycle_months,
    parse_dates,
    safe_name,
    save_uploaded_file,
)
from store import now_text


INSTRUMENT_COLUMNS = [
    "id",
    "management_no",
    "name",
    "serial_no",
    "cycle_text",
    "cycle_months",
    "location",
    "process",
    "department",
    "department_owner",
    "department_owner2",
    "qc_owner",
    "is_standard",
    "status",
    "remark",
    "history_card_updated",
    "correction_offset",
    "correction_factor",
    "correction_unit",
    "correction_note",
    "disposal_report_no",
    "disposal_report_file_path",
    "created_at",
    "updated_at",
]

RECORD_COLUMNS = [
    "id",
    "instrument_id",
    "calibration_type",
    "calibration_date",
    "next_due_date",
    "result",
    "certificate_no",
    "certificate_file_path",
    "measured_value",
    "corrected_value",
    "correction_snapshot",
    "note",
    "created_at",
]

CONTACT_COLUMNS = ["id", "department", "owner_name", "kakao_target", "phone", "note", "updated_at"]
IMPORT_LOG_COLUMNS = ["id", "source_path", "imported_at", "instrument_count", "record_count"]

SHEETS = {
    "instruments": INSTRUMENT_COLUMNS,
    "calibration_records": RECORD_COLUMNS,
    "department_contacts": CONTACT_COLUMNS,
    "import_log": IMPORT_LOG_COLUMNS,
}


def spreadsheet_id() -> str:
    value = get_secret("GOOGLE_SHEET_ID")
    if not value:
        raise RuntimeError("GOOGLE_SHEET_ID is not configured.")
    return value


def service():
    return sheets_service()


def init_db() -> None:
    svc = service()
    sid = spreadsheet_id()
    meta = svc.spreadsheets().get(spreadsheetId=sid).execute()
    existing = {sheet["properties"]["title"] for sheet in meta.get("sheets", [])}
    requests = [{"addSheet": {"properties": {"title": name}}} for name in SHEETS if name not in existing]
    if requests:
        svc.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": requests}).execute()
    for name, columns in SHEETS.items():
        values = get_values(name)
        if not values:
            write_values(name, [columns])
        elif values[0] != columns:
            rows = rows_to_dicts(name)
            write_table(name, pd.DataFrame(rows), columns)


def get_values(sheet_name: str) -> list[list[str]]:
    result = service().spreadsheets().values().get(spreadsheetId=spreadsheet_id(), range=f"{sheet_name}!A:ZZ").execute()
    return result.get("values", [])


def write_values(sheet_name: str, values: list[list[Any]]) -> None:
    svc = service()
    sid = spreadsheet_id()
    svc.spreadsheets().values().clear(spreadsheetId=sid, range=f"{sheet_name}!A:ZZ").execute()
    svc.spreadsheets().values().update(
        spreadsheetId=sid,
        range=f"{sheet_name}!A1",
        valueInputOption="RAW",
        body={"values": values},
    ).execute()


def rows_to_dicts(sheet_name: str) -> list[dict[str, Any]]:
    columns = SHEETS[sheet_name]
    values = get_values(sheet_name)
    if not values:
        return []
    header = values[0]
    rows = []
    for raw in values[1:]:
        padded = raw + [""] * (len(header) - len(raw))
        row = dict(zip(header, padded))
        rows.append({col: row.get(col, "") for col in columns})
    return rows


def table_df(sheet_name: str) -> pd.DataFrame:
    return pd.DataFrame(rows_to_dicts(sheet_name), columns=SHEETS[sheet_name])


def write_table(sheet_name: str, df: pd.DataFrame, columns: list[str] | None = None) -> None:
    columns = columns or SHEETS[sheet_name]
    if df.empty:
        write_values(sheet_name, [columns])
        return
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    out = out[columns].fillna("").astype(str)
    write_values(sheet_name, [columns] + out.values.tolist())


def next_id(sheet_name: str) -> int:
    df = table_df(sheet_name)
    if df.empty or "id" not in df:
        return 1
    ids = pd.to_numeric(df["id"], errors="coerce").dropna()
    return int(ids.max()) + 1 if not ids.empty else 1


def normalize_instruments(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    for col in ["id", "cycle_months", "is_standard"]:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    for col in ["correction_offset", "correction_factor"]:
        if col in df:
            default = 1 if col == "correction_factor" else 0
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(default)
    return df


def normalize_records(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    for col in ["id", "instrument_id"]:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    for col in ["measured_value", "corrected_value"]:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def instruments_raw_df() -> pd.DataFrame:
    return normalize_instruments(table_df("instruments"))


def records_raw_df() -> pd.DataFrame:
    return normalize_records(table_df("calibration_records"))


def latest_record_df() -> pd.DataFrame:
    instruments = instruments_raw_df()
    records = records_raw_df()
    if instruments.empty:
        return pd.DataFrame(columns=INSTRUMENT_COLUMNS)
    if records.empty:
        for col in [
            "last_record_id",
            "last_calibration_type",
            "last_calibration_date",
            "next_due_date",
            "last_result",
            "last_certificate_no",
            "last_certificate_file_path",
            "last_measured_value",
            "last_corrected_value",
        ]:
            instruments[col] = ""
        return instruments
    records["_due_sort"] = pd.to_datetime(records["next_due_date"], errors="coerce")
    records = records.sort_values(["instrument_id", "_due_sort", "id"], ascending=[True, False, False])
    latest = records.drop_duplicates("instrument_id")
    latest = latest.rename(
        columns={
            "id": "last_record_id",
            "calibration_type": "last_calibration_type",
            "calibration_date": "last_calibration_date",
            "result": "last_result",
            "certificate_no": "last_certificate_no",
            "certificate_file_path": "last_certificate_file_path",
            "measured_value": "last_measured_value",
            "corrected_value": "last_corrected_value",
        }
    )
    return instruments.merge(
        latest[
            [
                "instrument_id",
                "last_record_id",
                "last_calibration_type",
                "last_calibration_date",
                "next_due_date",
                "last_result",
                "last_certificate_no",
                "last_certificate_file_path",
                "last_measured_value",
                "last_corrected_value",
            ]
        ],
        how="left",
        left_on="id",
        right_on="instrument_id",
    ).drop(columns=["instrument_id"], errors="ignore")


def instruments_df(include_disposed: bool = True) -> pd.DataFrame:
    df = latest_record_df()
    if not include_disposed and not df.empty:
        df = df[df["status"] != "폐기"]
    return df


def calibration_history_df(instrument_id: int | None = None) -> pd.DataFrame:
    records = records_raw_df()
    instruments = instruments_raw_df()[["id", "management_no", "name"]] if not instruments_raw_df().empty else pd.DataFrame(columns=["id", "management_no", "name"])
    if records.empty:
        return pd.DataFrame(columns=["id", "instrument_id", "management_no", "name"] + RECORD_COLUMNS[2:])
    if instrument_id:
        records = records[records["instrument_id"] == int(instrument_id)]
    out = records.merge(instruments, how="left", left_on="instrument_id", right_on="id", suffixes=("", "_instrument"))
    if "id_instrument" in out:
        out = out.drop(columns=["id_instrument"])
    order = [
        "id",
        "instrument_id",
        "management_no",
        "name",
        "calibration_type",
        "calibration_date",
        "next_due_date",
        "result",
        "certificate_no",
        "certificate_file_path",
        "measured_value",
        "corrected_value",
        "note",
        "created_at",
    ]
    for col in order:
        if col not in out:
            out[col] = ""
    return out[order].sort_values(["next_due_date", "id"], ascending=[False, False])


def contacts_df() -> pd.DataFrame:
    return table_df("department_contacts")


def get_import_log() -> pd.DataFrame:
    return table_df("import_log").sort_values("id", ascending=False) if not table_df("import_log").empty else table_df("import_log")


def get_instrument(instrument_id: int) -> dict[str, Any] | None:
    df = instruments_raw_df()
    if df.empty:
        return None
    row = df[df["id"] == int(instrument_id)]
    return row.iloc[0].to_dict() if not row.empty else None


def upsert_instrument(data: dict[str, Any]) -> int:
    init_db()
    df = instruments_raw_df()
    ts = now_text()
    management_no = clean_text(data["management_no"])
    cycle_text = clean_text(data.get("cycle_text"))
    payload = {
        "management_no": management_no,
        "name": clean_text(data["name"]),
        "serial_no": clean_text(data.get("serial_no")),
        "cycle_text": cycle_text,
        "cycle_months": data.get("cycle_months") or parse_cycle_months(cycle_text) or "",
        "location": clean_text(data.get("location")),
        "process": clean_text(data.get("process")),
        "department": clean_text(data.get("department")) or derive_department(data.get("location", "")),
        "department_owner": clean_text(data.get("department_owner")),
        "department_owner2": clean_text(data.get("department_owner2")),
        "qc_owner": clean_text(data.get("qc_owner")),
        "is_standard": 1 if data.get("is_standard") else 0,
        "status": clean_text(data.get("status")) or "사용",
        "remark": clean_text(data.get("remark")),
        "history_card_updated": clean_text(data.get("history_card_updated")),
        "correction_offset": float(data.get("correction_offset") or 0),
        "correction_factor": float(data.get("correction_factor") or 1),
        "correction_unit": clean_text(data.get("correction_unit")),
        "correction_note": clean_text(data.get("correction_note")),
        "updated_at": ts,
    }
    if df.empty or management_no not in set(df["management_no"]):
        payload["id"] = next_id("instruments")
        payload["created_at"] = ts
        df = pd.concat([df, pd.DataFrame([payload])], ignore_index=True)
        write_table("instruments", df, INSTRUMENT_COLUMNS)
        return int(payload["id"])
    idx = df.index[df["management_no"] == management_no][0]
    for key, value in payload.items():
        df.at[idx, key] = value
    write_table("instruments", df, INSTRUMENT_COLUMNS)
    return int(df.at[idx, "id"])


def add_calibration_record(data: dict[str, Any]) -> None:
    init_db()
    df = records_raw_df()
    row = {
        "id": next_id("calibration_records"),
        "instrument_id": int(data["instrument_id"]),
        "calibration_type": clean_text(data.get("calibration_type")),
        "calibration_date": clean_text(data.get("calibration_date")),
        "next_due_date": clean_text(data.get("next_due_date")),
        "result": clean_text(data.get("result")),
        "certificate_no": clean_text(data.get("certificate_no")),
        "certificate_file_path": clean_text(data.get("certificate_file_path")),
        "measured_value": data.get("measured_value") if data.get("measured_value") is not None else "",
        "corrected_value": data.get("corrected_value") if data.get("corrected_value") is not None else "",
        "correction_snapshot": clean_text(data.get("correction_snapshot")),
        "note": clean_text(data.get("note")),
        "created_at": now_text(),
    }
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    write_table("calibration_records", df, RECORD_COLUMNS)


def update_instrument_master(instrument_id: int, data: dict[str, Any]) -> None:
    df = instruments_raw_df()
    idxs = df.index[df["id"] == int(instrument_id)]
    if idxs.empty:
        return
    idx = idxs[0]
    updates = {
        "name": clean_text(data.get("name")),
        "serial_no": clean_text(data.get("serial_no")),
        "cycle_text": clean_text(data.get("cycle_text")),
        "cycle_months": data.get("cycle_months") or parse_cycle_months(clean_text(data.get("cycle_text"))) or "",
        "location": clean_text(data.get("location")),
        "process": clean_text(data.get("process")),
        "department": clean_text(data.get("department")),
        "department_owner": clean_text(data.get("department_owner")),
        "department_owner2": clean_text(data.get("department_owner2")),
        "qc_owner": clean_text(data.get("qc_owner")),
        "is_standard": 1 if data.get("is_standard") else 0,
        "status": clean_text(data.get("status")) or "사용",
        "remark": clean_text(data.get("remark")),
        "updated_at": now_text(),
    }
    for key, value in updates.items():
        df.at[idx, key] = value
    write_table("instruments", df, INSTRUMENT_COLUMNS)


def update_correction(instrument_id: int, offset: float, factor: float, unit: str, note: str) -> None:
    df = instruments_raw_df()
    idxs = df.index[df["id"] == int(instrument_id)]
    if idxs.empty:
        return
    idx = idxs[0]
    df.at[idx, "correction_offset"] = offset
    df.at[idx, "correction_factor"] = factor
    df.at[idx, "correction_unit"] = clean_text(unit)
    df.at[idx, "correction_note"] = clean_text(note)
    df.at[idx, "updated_at"] = now_text()
    write_table("instruments", df, INSTRUMENT_COLUMNS)


def mark_disposed(instrument_id: int, note: str, report_no: str = "", report_file_path: str = "") -> None:
    df = instruments_raw_df()
    idxs = df.index[df["id"] == int(instrument_id)]
    if idxs.empty:
        return
    idx = idxs[0]
    existing = clean_text(df.at[idx, "remark"])
    note = clean_text(note)
    df.at[idx, "status"] = "폐기"
    if report_no:
        df.at[idx, "disposal_report_no"] = clean_text(report_no)
    if report_file_path:
        df.at[idx, "disposal_report_file_path"] = clean_text(report_file_path)
    if note:
        df.at[idx, "remark"] = f"{existing}\n{note}".strip() if existing else note
    df.at[idx, "updated_at"] = now_text()
    write_table("instruments", df, INSTRUMENT_COLUMNS)


def upsert_contact(department: str, owner_name: str, kakao_target: str, phone: str, note: str) -> None:
    init_db()
    df = table_df("department_contacts")
    payload = {
        "department": clean_text(department),
        "owner_name": clean_text(owner_name),
        "kakao_target": clean_text(kakao_target),
        "phone": clean_text(phone),
        "note": clean_text(note),
        "updated_at": now_text(),
    }
    if df.empty or payload["department"] not in set(df["department"]):
        payload["id"] = next_id("department_contacts")
        df = pd.concat([df, pd.DataFrame([payload])], ignore_index=True)
    else:
        idx = df.index[df["department"] == payload["department"]][0]
        for key, value in payload.items():
            df.at[idx, key] = value
    write_table("department_contacts", df, CONTACT_COLUMNS)


def dashboard_metrics() -> dict[str, int]:
    df = instruments_df(include_disposed=True)
    if df.empty:
        return {"total": 0, "active": 0, "disposed": 0, "overdue": 0, "due_90": 0}
    today = pd.Timestamp(date.today())
    due = pd.to_datetime(df["next_due_date"], errors="coerce")
    active = df["status"] != "폐기"
    return {
        "total": int(len(df)),
        "active": int(active.sum()),
        "disposed": int((df["status"] == "폐기").sum()),
        "overdue": int(((due < today) & active).sum()),
        "due_90": int(((due >= today) & (due <= today + pd.Timedelta(days=90)) & active).sum()),
    }


def due_items(days: int = 90, include_overdue: bool = True, calibration_filter: str = "전체") -> pd.DataFrame:
    df = instruments_df(include_disposed=False)
    if df.empty:
        return df
    if calibration_filter != "전체":
        df = df[df["last_calibration_type"].fillna("") == calibration_filter]
    today = pd.Timestamp(date.today())
    due = pd.to_datetime(df["next_due_date"], errors="coerce")
    mask = due <= today + pd.Timedelta(days=days)
    if not include_overdue:
        mask &= due >= today
    df = df[mask].copy()
    df["due_dt"] = due[mask]
    df["남은일수"] = (df["due_dt"] - today).dt.days
    return df.sort_values(["남은일수", "department", "management_no"])


def import_excel(path: str | Path, reset: bool = False) -> ImportSummary:
    init_db()
    if reset:
        for sheet_name, columns in SHEETS.items():
            write_values(sheet_name, [columns])
    source = Path(path)
    wb = openpyxl.load_workbook(source, data_only=True)
    instrument_count = 0
    record_count = 0
    disposed_count = 0
    ws = wb["검교정 내역-2026.05.15"] if "검교정 내역-2026.05.15" in wb.sheetnames else wb.worksheets[0]
    for row in ws.iter_rows(min_row=4, values_only=True):
        management_no = clean_text(row[0] if len(row) > 0 else "")
        name = clean_text(row[1] if len(row) > 1 else "")
        if not management_no or not name:
            continue
        remark = clean_text(row[12] if len(row) > 12 else "")
        disposal_text = " ".join([remark, clean_text(row[9] if len(row) > 9 else ""), clean_text(row[10] if len(row) > 10 else "")])
        status = "폐기" if "폐기" in disposal_text else "사용"
        disposed_count += 1 if status == "폐기" else 0
        instrument_id = upsert_instrument(
            {
                "management_no": management_no,
                "name": name,
                "serial_no": clean_text(row[2] if len(row) > 2 else ""),
                "cycle_text": clean_text(row[3] if len(row) > 3 else ""),
                "location": clean_text(row[4] if len(row) > 4 else ""),
                "process": clean_text(row[5] if len(row) > 5 else ""),
                "department": clean_text(row[6] if len(row) > 6 else "") or derive_department(row[4] if len(row) > 4 else ""),
                "department_owner": clean_text(row[7] if len(row) > 7 else ""),
                "department_owner2": clean_text(row[8] if len(row) > 8 else ""),
                "is_standard": "표준품" in management_no or "표준품" in remark,
                "status": status,
                "remark": remark,
                "history_card_updated": clean_text(row[13] if len(row) > 13 else ""),
            }
        )
        instrument_count += 1
        for calibration_type, value in (("내부", row[9] if len(row) > 9 else None), ("외부", row[10] if len(row) > 10 else None)):
            text = clean_text(value)
            if not text or text.upper() == "N/A" or "폐기" in text:
                continue
            parsed = parse_dates(text)
            add_calibration_record(
                {
                    "instrument_id": instrument_id,
                    "calibration_type": calibration_type,
                    "calibration_date": parsed[0].isoformat() if parsed else "",
                    "next_due_date": parsed[-1].isoformat() if len(parsed) > 1 else "",
                    "result": "기존 대장",
                    "note": text,
                }
            )
            record_count += 1
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            sterilizer_no = clean_text(row[0] if len(row) > 0 else "")
            name = clean_text(row[1] if len(row) > 1 else "")
            serial_no = clean_text(row[2] if len(row) > 2 else "")
            sensor_no = clean_text(row[3] if len(row) > 3 else "")
            if not name or not serial_no:
                continue
            instrument_id = upsert_instrument(
                {
                    "management_no": f"멸균-{sterilizer_no}-{serial_no}".replace(" ", ""),
                    "name": name,
                    "serial_no": serial_no,
                    "cycle_text": "12개월",
                    "location": sterilizer_no,
                    "process": "멸균기 부착",
                    "department": "멸균",
                    "status": "사용",
                    "remark": f"멸균기 센서 No: {sensor_no}" if sensor_no else "",
                }
            )
            instrument_count += 1
            cal_dates = parse_dates(row[4] if len(row) > 4 else "")
            due_dates = parse_dates(row[5] if len(row) > 5 else "")
            add_calibration_record(
                {
                    "instrument_id": instrument_id,
                    "calibration_type": "외부",
                    "calibration_date": cal_dates[0].isoformat() if cal_dates else "",
                    "next_due_date": due_dates[0].isoformat() if due_dates else "",
                    "result": "기존 대장",
                    "note": "멸균기 부착 계측기",
                }
            )
            record_count += 1
    log = table_df("import_log")
    log_row = {
        "id": next_id("import_log"),
        "source_path": str(source),
        "imported_at": now_text(),
        "instrument_count": instrument_count,
        "record_count": record_count,
    }
    write_table("import_log", pd.concat([log, pd.DataFrame([log_row])], ignore_index=True), IMPORT_LOG_COLUMNS)
    return ImportSummary(instrument_count, record_count, disposed_count, str(source))
